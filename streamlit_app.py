"""
Streamlit web app for the Automatic Exam Seating Arrangement System.

Tab 1 – Convert Result Workbook
    Upload your existing result Excel → one-click auto-extract students + rooms
    → preview data → download ready-to-use input files

Tab 2 – Generate Seating Plan
    Upload (or use auto-extracted) students + rooms files
    → configure options → generate seating
    → download SIRT-formatted outputs (Main Seating, Debarred List, Attendance Sheets, Excel)
"""
from __future__ import annotations

import io
from pathlib import Path
from uuid import uuid4

import pandas as pd
import streamlit as st

from convert_result_to_students import convert_result_file
from exam_seating.io_utils import DataValidationError
from exam_seating.service import SeatingConfig, run_seating_system
from exam_seating.sirt_exporters import (
    export_sirt_attendance_sheets_excel,
    export_sirt_debarred_excel,
    export_sirt_main_seating_excel,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _save_uploaded(uploaded_file, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / uploaded_file.name
    path.write_bytes(uploaded_file.getbuffer())
    return path


def _df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


def _make_students_template() -> bytes:
    df = pd.DataFrame(
        [
            {"Roll Number": "0133CI231002", "Name": "AARBI DHAKAD",       "Attendance Percentage": 82},
            {"Roll Number": "0133CI231008", "Name": "ABHISHEK KUMAR",     "Attendance Percentage": 61},
            {"Roll Number": "0133CI231014", "Name": "ADITYA KUMAR SHAH",  "Attendance Percentage": 75},
            {"Roll Number": "0133CI231024", "Name": "AMANJEET SINGH",     "Attendance Percentage": 38},
            {"Roll Number": "0133CI231035", "Name": "ANJALI SHRIVASTAVA", "Attendance Percentage": 55},
        ]
    )
    return _df_to_excel_bytes(df)


def _make_rooms_template() -> bytes:
    df = pd.DataFrame(
        [
            {"Room Number": "F-307", "Capacity": 40},
            {"Room Number": "F-308", "Capacity": 40},
            {"Room Number": "F-309", "Capacity": 40},
            {"Room Number": "F-310", "Capacity": 30},
        ]
    )
    return _df_to_excel_bytes(df)


def _make_room_exam_config_template() -> bytes:
    """Template for room exam configuration with room, capacity, subject, and date."""
    df = pd.DataFrame(
        [
            {"Room Number": "F-307", "Capacity": 40, "Subject": "Computer Networks", "Date": "2026-06-05"},
            {"Room Number": "F-308", "Capacity": 40, "Subject": "Database Systems",  "Date": "2026-06-05"},
            {"Room Number": "F-309", "Capacity": 40, "Subject": "Web Development",   "Date": "2026-06-06"},
            {"Room Number": "F-310", "Capacity": 30, "Subject": "Data Structures",   "Date": "2026-06-06"},
        ]
    )
    return _df_to_excel_bytes(df)


def _make_sample_result_workbook() -> bytes:
    """Generate a realistic SIRT-style result workbook with header rows, 3 branch sheets,
    and a Main Seating sheet — ready to upload in Step 1."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        wb = writer.book

        # ── Styles ────────────────────────────────────────────────────────────
        hdr_fmt = wb.add_format({
            "bold": True, "align": "center", "valign": "vcenter",
            "font_size": 12, "border": 1,
        })
        col_fmt = wb.add_format({
            "bold": True, "align": "center", "valign": "vcenter",
            "border": 1, "bg_color": "#D9E1F2",
        })
        cell_fmt = wb.add_format({"border": 1, "align": "center", "valign": "vcenter"})

        # ── Branch sheets ─────────────────────────────────────────────────────
        INSTITUTE = "SAGAR INSTITUTE OF RESEARCH & TECHNOLOGY"
        DEPT      = "DEPARTMENT OF COMPUTER SCIENCE & INFORMATION TECHNOLOGY (CSIT)"
        EXAM      = "I MID SEM EXAMINATION (JAN–JUN 2026)"

        branches = {
            "6th A": [
                ("0133CI231002", "AARBI DHAKAD",           85),
                ("0133CI231008", "ABHISHEK KUMAR",          62),
                ("0133CI231014", "ADITYA KUMAR SHAH",       78),
                ("0133CI231024", "AMANJEET SINGH",          35),
                ("0133CI231035", "ANJALI SHRIVASTAVA",      55),
                ("0133CI231041", "ANKIT PATEL",             90),
                ("0133CI231047", "ANSHUL VERMA",            48),
                ("0133CI231053", "ARPIT MISHRA",            72),
                ("0133CI231059", "ATUL SHARMA",             66),
                ("0133CI231065", "BHAVESH TIWARI",          80),
            ],
            "6th B": [
                ("0133ME231001", "AAKASH SINGH",            70),
                ("0133ME231007", "AASTHA GUPTA",            55),
                ("0133ME231013", "ABHAY KUMAR",             38),
                ("0133ME231019", "ABHINAV RAI",             85),
                ("0133ME231025", "ADITI JAIN",              60),
                ("0133ME231031", "ADITYA THAKUR",           77),
                ("0133ME231037", "AJAY BAGHEL",             92),
                ("0133ME231043", "AKASH CHOUHAN",           44),
                ("0133ME231049", "AKSHAY KORI",             68),
                ("0133ME231055", "ALKA PATEL",              73),
            ],
            "6th C": [
                ("0133EC231003", "AKASH SAHU",              81),
                ("0133EC231009", "ALISHA KHAN",             59),
                ("0133EC231015", "AMANDEEP KHATRI",         36),
                ("0133EC231021", "AMISHA DUBEY",            74),
                ("0133EC231027", "AMRIT PAL",               88),
                ("0133EC231033", "ANANYA SHUKLA",           50),
                ("0133EC231039", "ANKITA YADAV",            67),
                ("0133EC231045", "ANSHIKA TRIPATHI",        41),
                ("0133EC231051", "ANURAG TIWARI",           95),
                ("0133EC231057", "ARJUN SINGH",             63),
            ],
        }

        for sheet_name, students in branches.items():
            ws = wb.add_worksheet(sheet_name)
            ws.set_column(0, 0, 6)
            ws.set_column(1, 1, 22)
            ws.set_column(2, 2, 32)
            ws.set_column(3, 3, 22)

            # Header rows (institute branding)
            ws.set_row(0, 20)
            ws.set_row(1, 18)
            ws.set_row(2, 16)
            ws.merge_range("A1:D1", INSTITUTE, hdr_fmt)
            ws.merge_range("A2:D2", DEPT,      hdr_fmt)
            ws.merge_range("A3:D3", EXAM,      hdr_fmt)
            ws.merge_range("A4:D4", f"Subject: Computer Networks    {sheet_name}", hdr_fmt)

            # Column headers at row 5 (0-indexed: row 4)
            for col, label in enumerate(["S.No", "Board EnrollNo", "Student Name", "Attendance %"]):
                ws.write(4, col, label, col_fmt)

            # Student data
            for i, (roll, name, att) in enumerate(students):
                ws.write(5 + i, 0, i + 1,   cell_fmt)
                ws.write(5 + i, 1, roll,     cell_fmt)
                ws.write(5 + i, 2, name,     cell_fmt)
                ws.write(5 + i, 3, f"{att}%", cell_fmt)

        # ── Main Seating sheet ────────────────────────────────────────────────
        ms = wb.add_worksheet("Main Seating")
        ms.set_column(0, 0, 8)
        ms.set_column(1, 1, 16)
        ms.set_column(2, 2, 12)
        ms.set_column(3, 3, 32)
        ms.set_column(4, 4, 26)
        ms.set_column(5, 5, 14)

        ms.merge_range("A1:F1", INSTITUTE, hdr_fmt)
        ms.merge_range("A2:F2", EXAM, hdr_fmt)
        ms.merge_range("A3:F3", "MAIN SEATING ARRANGEMENT", hdr_fmt)

        ms_cols = ["S. No.", "BRANCH", "Section", "Roll Number", "No. of Students Appeared", "Class Room"]
        for col, label in enumerate(ms_cols):
            ms.write(3, col, label, col_fmt)

        room_rows = [
            (1, "CSIT",  "A", "0133CI231002, 0133CI231008, 0133CI231014, 0133CI231024, 0133CI231035, 0133CI231041, 0133CI231047, 0133CI231053, 0133CI231059, 0133CI231065", 10, "F-307"),
            (2, "ME",    "B", "0133ME231001, 0133ME231007, 0133ME231013, 0133ME231019, 0133ME231025, 0133ME231031, 0133ME231037, 0133ME231043, 0133ME231049, 0133ME231055", 10, "F-308"),
            (3, "EC",    "C", "0133EC231003, 0133EC231009, 0133EC231015, 0133EC231021, 0133EC231027, 0133EC231033, 0133EC231039, 0133EC231045, 0133EC231051, 0133EC231057", 10, "F-309"),
        ]
        for r, row in enumerate(room_rows):
            for c, val in enumerate(row):
                ms.write(4 + r, c, val, cell_fmt)

        # ── Debarred sheet (bonus) ────────────────────────────────────────────
        db = wb.add_worksheet("Debarred")
        db.merge_range("A1:E1", INSTITUTE, hdr_fmt)
        db.merge_range("A2:E2", "Debarred Students List (Attendance < 40%)", hdr_fmt)
        for col, label in enumerate(["S.No", "Board EnrollNo", "Student Name", "Attendance %", "Section"]):
            db.write(2, col, label, col_fmt)
        debarred = [
            (1, "0133CI231024", "AMANJEET SINGH",    "35%", "A"),
            (2, "0133ME231013", "ABHAY KUMAR",        "38%", "B"),
            (3, "0133EC231015", "AMANDEEP KHATRI",    "36%", "C"),
            (4, "0133ME231043", "AKASH CHOUHAN",      "44%", "B"),
            (5, "0133EC231045", "ANSHIKA TRIPATHI",   "41%", "C"),
        ]
        for r, row in enumerate(debarred):
            for c, val in enumerate(row):
                db.write(3 + r, c, val, cell_fmt)

    return buf.getvalue()


def _merge_workbooks(paths: list, labels: list[str]) -> bytes:
    """Merge multiple xlsx files into one workbook, prefixing sheet names with a label."""
    import io as _io
    from openpyxl import load_workbook
    from openpyxl import Workbook as _WB
    from copy import copy

    merged = _WB()
    merged.remove(merged.active)  # remove default empty sheet

    for path, label in zip(paths, labels):
        wb = load_workbook(path)
        for src_ws in wb.worksheets:
            raw_name = f"{label}-{src_ws.title}" if len(wb.worksheets) > 1 else label
            safe = "".join(c for c in raw_name if c not in r'\/*?:[]')[:31]
            dst_ws = merged.create_sheet(title=safe)

            # Copy column dimensions
            for col, dim in src_ws.column_dimensions.items():
                dst_ws.column_dimensions[col].width = dim.width

            # Copy row dimensions
            for row, dim in src_ws.row_dimensions.items():
                dst_ws.row_dimensions[row].height = dim.height

            # Copy merged cell ranges
            for merge in src_ws.merged_cells.ranges:
                dst_ws.merge_cells(str(merge))

            # Copy cells (value + style)
            for row in src_ws.iter_rows():
                for cell in row:
                    dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        dst_cell.font      = copy(cell.font)
                        dst_cell.fill      = copy(cell.fill)
                        dst_cell.border    = copy(cell.border)
                        dst_cell.alignment = copy(cell.alignment)

    buf = _io.BytesIO()
    merged.save(buf)
    return buf.getvalue()


def _sidebar_institute_info() -> dict:
    st.sidebar.header("Institute Info")
    institute = st.sidebar.text_input("Institute Name", "SAGAR INSTITUTE OF RESEARCH & TECHNOLOGY")
    department = st.sidebar.text_input("Department", "DEPARTMENT OF CSIT")
    exam_title = st.sidebar.text_input("Exam Title", "I MID SEM EXAMINATION (JAN-JUN 2026)")
    semester = st.sidebar.text_input("Semester", "6th Sem")

    return {"institute": institute, "department": department,
            "exam_title": exam_title, "semester": semester}


def tab_convert() -> None:
    st.header("📂 Step 1 — Convert Result Workbook")
    st.markdown(
        """
        Upload your **SIRT result Excel file** (.xlsx / .xls).  
        The app will automatically detect every branch sheet, extract student roll numbers,
        names, and attendance, and pull exam room data from the **Main Seating** sheet —
        all in one click.  
        The extracted files are then passed directly to **Step 2** so you never need to
        prepare input files by hand.
        """
    )

    # ── How it works ──────────────────────────────────────────────────────────
    with st.expander("ℹ️  How does this work?", expanded=False):
        st.markdown(
            """
            | Stage | What happens |
            |---|---|
            | **Upload** | You upload the result Excel workbook (the same file shared by the exam section). |
            | **Sheet detection** | All branch/section sheets (e.g. *6th A*, *6th B*) are listed. Utility sheets like *Main Seating* and *Debarred* are excluded automatically. |
            | **Extraction** | Each sheet is scanned for the student table. Roll number, name, and attendance % are normalised. Section and branch are inferred from the sheet name. |
            | **Rooms** | Room numbers and capacities are read from the *Main Seating* sheet (if present). |
            | **Download / Continue** | Preview both tables, then either download them as `.xlsx` files or go straight to Step 2 — the data is already loaded. |
            """
        )
        st.markdown("**Want to try it with a sample file first?**")
        st.download_button(
            label="⬇ Download sample_result_workbook.xlsx",
            data=_make_sample_result_workbook(),
            file_name="sample_result_workbook.xlsx",
            mime=XLSX_MIME,
            help="A realistic SIRT-format workbook with 3 branch sheets, a Main Seating sheet, and a Debarred sheet.",
            key="sample_wb_expander",
        )

    # ── Sample templates ──────────────────────────────────────────────────────
    with st.expander("📥  Don't have a result workbook? Use these manual input templates", expanded=False):
        st.markdown("If you want to enter student and room data by hand instead of uploading a result workbook, download these starter files, fill them in, and upload them in **Step 2**.")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(
                "**students_template.xlsx**  \n"
                "Columns: `Roll Number` · `Name` · `Attendance Percentage`"
            )
            st.download_button(
                "⬇ Download students_template.xlsx",
                data=_make_students_template(),
                file_name="students_template.xlsx",
                mime=XLSX_MIME,
                key="tmpl_students_tab1",
            )
        with c2:
            st.markdown(
                "**rooms_template.xlsx**  \n"
                "Columns: `Room Number` · `Capacity`"
            )
            st.download_button(
                "⬇ Download rooms_template.xlsx",
                data=_make_rooms_template(),
                file_name="rooms_template.xlsx",
                mime=XLSX_MIME,
                key="tmpl_rooms_tab1",
            )

    st.divider()

    # ── Upload ─────────────────────────────────────────────────────────────────
    st.subheader("📤 Upload your result workbook")
    st.caption("New here? Download the sample workbook below, then upload it to see how extraction works.")
    st.download_button(
        label="⬇ Download sample_result_workbook.xlsx  (try with this first)",
        data=_make_sample_result_workbook(),
        file_name="sample_result_workbook.xlsx",
        mime=XLSX_MIME,
        help="Contains 3 branch sheets (6th A/B/C), Main Seating, and Debarred — mirrors a real SIRT result file.",
        key="sample_wb_upload_section",
    )
    result_file = st.file_uploader(
        "Accepted formats: .xlsx, .xls",
        type=["xlsx", "xls"],
        key="result_upload",
        help="Upload the Excel result file received from the exam section.",
    )

    if result_file is None:
        st.info("⬆️  Upload a result workbook above, then click **Extract Students & Rooms**.")
        return

    run_id = uuid4().hex[:8]
    tmp_dir = Path("web_runs") / run_id / "input"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    result_path = tmp_dir / result_file.name
    result_path.write_bytes(result_file.getbuffer())

    # ── Sheet selector ─────────────────────────────────────────────────────────
    xl = pd.ExcelFile(result_path)
    all_sheets = xl.sheet_names
    default_selected = [
        s for s in all_sheets
        if "debarred" not in s.lower()
        and s.lower() not in {"sheet1", "over all result", "exam duty chart", "main seating"}
    ]
    st.subheader("📋 Select student sheets")
    st.caption(
        "All sheets are listed below. Deselect any that do not contain student data "
        "(e.g. summary sheets, duty charts)."
    )
    selected_sheets = st.multiselect(
        "Sheets to include",
        options=all_sheets,
        default=default_selected,
        help="Only selected sheets will be scanned for student records.",
    )
    if not selected_sheets:
        st.warning("Select at least one sheet before extracting.")
        return

    st.divider()

    # ── Extract button ─────────────────────────────────────────────────────────
    if st.button("⚙️  Extract Students & Rooms", type="primary", use_container_width=True):
        with st.spinner("Reading workbook and extracting data…"):
            try:
                students, rooms = convert_result_file(
                    result_path,
                    sheets=selected_sheets,
                    skip_debarred=False,
                )
            except Exception as exc:
                st.error(f"❌ Extraction failed: {exc}")
                return

        if students.empty:
            st.error("No student records found in the selected sheets. Check that the sheets contain a 'Board EnrollNo' or 'Roll Number' column.")
            return

        # ── Results ────────────────────────────────────────────────────────────
        m1, m2, m3 = st.columns(3)
        m1.metric("Total Students", len(students))
        m2.metric("Rooms Found", len(rooms))
        m3.metric("Sections", students["Section"].nunique() if "Section" in students.columns else "—")

        st.success("✅ Extraction complete! Preview your data and download below, or switch to **Step 2** to generate the seating plan right away.")

        col1, col2 = st.columns(2)
        with col1:
            st.subheader("👨‍🎓 Students")
            st.dataframe(students.head(20), use_container_width=True)
            if len(students) > 20:
                st.caption(f"Showing first 20 of {len(students)} rows.")
            st.download_button(
                label=f"⬇ Download students_extracted.xlsx  ({len(students)} rows)",
                data=_df_to_excel_bytes(students),
                file_name="students_extracted.xlsx",
                mime=XLSX_MIME,
                use_container_width=True,
            )
        with col2:
            st.subheader("🏫 Rooms")
            if rooms.empty:
                st.warning("No rooms found in the workbook. You can upload a rooms file manually in Step 2.")
            else:
                st.dataframe(rooms, use_container_width=True)
                st.download_button(
                    label=f"⬇ Download rooms_extracted.xlsx  ({len(rooms)} rows)",
                    data=_df_to_excel_bytes(rooms),
                    file_name="rooms_extracted.xlsx",
                    mime=XLSX_MIME,
                    use_container_width=True,
                )

        st.session_state["extracted_students"] = students
        st.session_state["extracted_rooms"] = rooms
        st.info("💡 Data has been loaded into memory. Switch to **Step 2 — Generate Seating Plan** to continue.")


def tab_generate(info: dict) -> None:
    st.header("Generate Seating Plan")

    # ── Session state ──────────────────────────────────────────────────────────
    if "rooms_with_info" not in st.session_state:
        st.session_state["rooms_with_info"] = []
    if "exam_subjects" not in st.session_state:
        st.session_state["exam_subjects"] = []

    # ── 1. Upload Students Excel ───────────────────────────────────────────────
    st.subheader("1. Upload Students Excel")
    st.caption(
        "Your file must have these columns: **Roll Number** · **Name** · **Attendance Percentage**"
    )
    col_tmpl, _ = st.columns([1, 3])
    col_tmpl.download_button(
        "⬇ Download template",
        data=_make_students_template(),
        file_name="students_template.xlsx",
        mime=XLSX_MIME,
        key="tmpl_students_gen",
    )
    students_file = st.file_uploader(
        "Upload students Excel (.xlsx / .xls / .csv)",
        type=["xlsx", "xls", "csv"],
        key="students_upload_gen",
    )

    students_df = pd.DataFrame()
    if students_file is not None:
        try:
            if students_file.name.endswith(".csv"):
                students_df = pd.read_csv(students_file)
            else:
                # Read ALL sheets and combine them
                xl = pd.ExcelFile(students_file)
                sheet_frames = []
                required = {"Roll Number", "Name", "Attendance Percentage"}
                skipped_sheets = []
                multi_sheet = len(xl.sheet_names) > 1
                for sheet in xl.sheet_names:
                    df_sheet = pd.read_excel(xl, sheet_name=sheet)
                    df_sheet.columns = [c.strip() for c in df_sheet.columns]
                    if required.issubset(set(df_sheet.columns)):
                        # If file has multiple sheets and no Section column,
                        # derive Section (and Branch) from the sheet name
                        if multi_sheet and "Section" not in df_sheet.columns:
                            import re as _re
                            m = _re.search(r"\b([A-Z])\s*$", sheet.strip())
                            df_sheet["Section"] = m.group(1) if m else sheet.strip()
                        if multi_sheet and "Branch" not in df_sheet.columns:
                            # Use everything before the trailing letter as branch, default CSIT
                            branch_part = _re.sub(r"\s*[A-Z]\s*$", "", sheet.strip()).strip()
                            df_sheet["Branch"] = branch_part if branch_part else "CSIT"
                        sheet_frames.append(df_sheet)
                    else:
                        skipped_sheets.append(sheet)
                if not sheet_frames:
                    st.error(
                        f"None of the {len(xl.sheet_names)} sheet(s) contain the required columns: "
                        f"Roll Number, Name, Attendance Percentage."
                    )
                    students_df = pd.DataFrame()
                else:
                    students_df = pd.concat(sheet_frames, ignore_index=True)
                    if skipped_sheets:
                        st.caption(f"ℹ️ Skipped sheets (missing required columns): {', '.join(skipped_sheets)}")
            # Normalise column names (for CSV path)
            if not students_df.empty:
                students_df.columns = [c.strip() for c in students_df.columns]
            required = {"Roll Number", "Name", "Attendance Percentage"}
            missing = required - set(students_df.columns)
            if not students_df.empty and missing:
                st.error(f"Missing columns in uploaded file: {', '.join(missing)}")
                students_df = pd.DataFrame()
            elif not students_df.empty:
                students_df = students_df.drop_duplicates(subset=["Roll Number"], keep="first")
                students_df = students_df.sort_values(by=["Roll Number"], kind="mergesort").reset_index(drop=True)
                sections = students_df["Section"].nunique() if "Section" in students_df.columns else None
                info_msg = f"Loaded **{len(students_df)}** students"
                if sections:
                    info_msg += f" across **{sections}** section(s)"
                st.success(info_msg + ".")
                if "Section" in students_df.columns:
                    for section, grp in students_df.groupby("Section", sort=True):
                        branch = grp["Branch"].iloc[0] if "Branch" in grp.columns else ""
                        label = f"{branch}-{section}" if branch else f"Section {section}"
                        with st.expander(f"📋 {label}  ({len(grp)} students)", expanded=False):
                            st.dataframe(
                                grp[["Roll Number", "Name", "Attendance Percentage"]].reset_index(drop=True),
                                use_container_width=True,
                                hide_index=True,
                            )
                else:
                    st.dataframe(students_df.head(10), use_container_width=True, hide_index=True)
                    if len(students_df) > 10:
                        st.caption(f"Showing first 10 of {len(students_df)} rows.")
        except Exception as exc:
            st.error(f"Could not read file: {exc}")

    st.divider()

    # ── 2. Add Rooms ──────────────────────────────────────────────────────────
    st.subheader("2. Add Rooms")

    # ── 2a. Exam Timetable (subjects & dates for attendance sheet) ────────────
    st.markdown("**Exam Timetable** *(subjects & dates — appear as column headers in attendance sheet)*")
    with st.form("add_subject_form", clear_on_submit=True):
        ts1, ts2, ts_btn = st.columns([2, 1.5, 0.8])
        subj_in  = ts1.text_input("Subject Code / Name", placeholder="CSIT-401 (M-III)")
        sdate_in = ts2.date_input("Exam Date", key="subj_date_input")
        add_subj = ts_btn.form_submit_button("➕ Add")
        if add_subj:
            if not subj_in:
                st.error("Subject is required.")
            else:
                st.session_state["exam_subjects"].append({
                    "subject": subj_in,
                    "date": sdate_in.strftime("%d-%b-%y"),
                })

    if st.session_state["exam_subjects"]:
        subj_display = pd.DataFrame(st.session_state["exam_subjects"])
        subj_display.index = range(1, len(subj_display) + 1)
        sc1, sc2 = st.columns([0.95, 0.05])
        sc1.dataframe(subj_display, use_container_width=True)
        if sc2.button("🗑️", key="del_subjects", help="Clear all subjects"):
            st.session_state["exam_subjects"] = []
            st.rerun()

    st.divider()

    # ── 2b. Room Number & Capacity ────────────────────────────────────────────
    st.markdown("**Room Number & Capacity**")
    with st.form("add_room_form", clear_on_submit=True):
        c1, c2, c_btn = st.columns([2, 1.5, 0.8])
        room_no_in  = c1.text_input("Room Number", placeholder="F-307")
        capacity_in = c2.number_input("Capacity", 1, 500, 40, 1)
        add_room    = c_btn.form_submit_button("➕ Add")

        if add_room:
            if not room_no_in:
                st.error("Room Number is required.")
            else:
                existing = {r["Room Number"] for r in st.session_state["rooms_with_info"]}
                if room_no_in in existing:
                    st.error(f"Room {room_no_in} already added.")
                else:
                    st.session_state["rooms_with_info"].append({
                        "Room Number": room_no_in,
                        "Capacity": int(capacity_in),
                    })

    if st.session_state["rooms_with_info"]:
        rooms_display = pd.DataFrame(st.session_state["rooms_with_info"])
        col_tbl, col_del = st.columns([0.95, 0.05])
        col_tbl.dataframe(rooms_display, use_container_width=True, hide_index=True)
        if col_del.button("🗑️", key="del_rooms", help="Clear all rooms"):
            st.session_state["rooms_with_info"] = []
            st.rerun()

    st.divider()

    # ── 3. Options ────────────────────────────────────────────────────────────
    st.subheader("3. Options")
    o1, o2, o3, o4 = st.columns(4)
    attendance_cutoff = o1.number_input("Attendance cutoff (%)", 0.0, 100.0, 40.0, 1.0)
    shuffle    = o2.checkbox("Shuffle students")
    seed       = o3.number_input("Seed (if shuffle)", 0, 99999, 42, 1)
    alt_seats  = o4.checkbox("Alternate seats")

    # ── Live eligibility & room requirement summary ───────────────────────────
    if not students_df.empty:
        eligible_count   = int((pd.to_numeric(students_df["Attendance Percentage"], errors="coerce") >= attendance_cutoff).sum())
        debarred_count   = len(students_df) - eligible_count
        total_capacity   = sum(r["Capacity"] for r in st.session_state["rooms_with_info"])
        seats_needed     = eligible_count

        sm1, sm2, sm3 = st.columns(3)
        sm1.metric("Eligible Students", eligible_count)
        sm2.metric("Debarred Students", debarred_count)
        sm3.metric("Total Room Capacity", total_capacity if total_capacity else "—")

        if st.session_state["rooms_with_info"]:
            spare = total_capacity - seats_needed
            if spare >= 0:
                st.success(f"✅ Rooms sufficient — {seats_needed} students fit with **{spare} seats spare**.")
                rooms_ok = True
            else:
                st.error(f"❌ Short by **{abs(spare)} seats** — add more rooms before generating.")
                rooms_ok = False
        else:
            st.info(f"ℹ️ Add rooms above — you need capacity for **{seats_needed}** eligible students.")
            rooms_ok = False
    else:
        rooms_ok = False

    st.divider()

    if not st.button("Generate Seating Plan", type="primary", use_container_width=True, disabled=not rooms_ok):
        return

    # ── Validation ────────────────────────────────────────────────────────────
    if students_df.empty:
        st.error("Upload a valid students file first.")
        return
    if not st.session_state["rooms_with_info"]:
        st.error("Add at least one room before generating.")
        return

    rooms_with_info = st.session_state["rooms_with_info"]
    rooms_df = pd.DataFrame([{"Room Number": r["Room Number"], "Capacity": r["Capacity"]}
                              for r in rooms_with_info])
    exam_subjects = st.session_state["exam_subjects"]

    run_id = uuid4().hex[:8]
    base_dir   = Path("web_runs") / run_id
    input_dir  = base_dir / "input"
    output_dir = base_dir / "output"
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        student_path = input_dir / "students.xlsx"
        room_path    = input_dir / "rooms.xlsx"
        students_df.to_excel(student_path, index=False)
        rooms_df.to_excel(room_path, index=False)

        config = SeatingConfig(
            students_file=str(student_path),
            rooms_file=str(room_path),
            attendance_cutoff=float(attendance_cutoff),
            output_dir=str(output_dir),
            shuffle_students=shuffle,
            random_seed=int(seed) if shuffle else None,
            alternate_seats=alt_seats,
            export_pdf_file=False,
        )
        result = run_seating_system(config)

    except (DataValidationError, FileNotFoundError) as exc:
        st.error(str(exc))
        return
    except Exception as exc:
        st.error(f"Unexpected error: {exc}")
        return

    st.success("Seating plan generated!")

    m = st.columns(4)
    m[0].metric("Total Students",          result.summary.total_students)
    m[1].metric("Eligible",                result.summary.eligible_students)
    m[2].metric("Allocated",               result.summary.allocated_students)
    m[3].metric("Debarred (Not Eligible)", result.summary.not_eligible_students)

    # Re-attach Section/Branch columns if present
    raw_students = pd.read_excel(student_path)
    has_section = "Section" in raw_students.columns
    has_branch  = "Branch"  in raw_students.columns
    extra_cols  = (["Section"] if has_section else []) + (["Branch"] if has_branch else [])

    alloc_sheets = pd.read_excel(output_dir / "exam_seating_output.xlsx", sheet_name=None)
    room_allocations_enriched: dict[str, pd.DataFrame] = {}
    for sheet_name, df in alloc_sheets.items():
        if not sheet_name.startswith("Room_"):
            continue
        room_no = sheet_name[len("Room_"):]
        if extra_cols:
            lookup = raw_students[["Roll Number"] + extra_cols].copy()
            lookup["Roll Number"] = lookup["Roll Number"].astype(str).str.strip()
            df["Roll Number"]     = df["Roll Number"].astype(str).str.strip()
            df = df.merge(lookup, on="Roll Number", how="left")
        room_allocations_enriched[room_no] = df

    not_eligible_df = pd.read_excel(output_dir / "exam_seating_output.xlsx", sheet_name="Not_Eligible")
    if extra_cols:
        lookup = raw_students[["Roll Number"] + extra_cols].copy()
        lookup["Roll Number"]            = lookup["Roll Number"].astype(str).str.strip()
        not_eligible_df["Roll Number"]   = not_eligible_df["Roll Number"].astype(str).str.strip()
        not_eligible_df = not_eligible_df.merge(lookup, on="Roll Number", how="left")

    sirt_seating_path  = export_sirt_main_seating_excel(
        room_allocations_enriched, output_dir,
        institute=info["institute"], department=info["department"],
        exam_title=info["exam_title"], semester=info["semester"],
    )
    sirt_debarred_path = export_sirt_debarred_excel(
        not_eligible_df, output_dir,
        institute=info["institute"], department=info["department"],
        exam_title=info["exam_title"], semester=info["semester"],
        attendance_cutoff=float(attendance_cutoff),
    )
    # Build eligible students df with Branch/Section for section-wise attendance sheets
    eligible_for_att = raw_students.copy()
    eligible_for_att["Attendance Percentage"] = pd.to_numeric(
        eligible_for_att["Attendance Percentage"], errors="coerce"
    )
    eligible_for_att = eligible_for_att[
        eligible_for_att["Attendance Percentage"] >= float(attendance_cutoff)
    ].reset_index(drop=True)

    sirt_att_path = export_sirt_attendance_sheets_excel(
        eligible_for_att, output_dir,
        subjects=exam_subjects,
        institute=info["institute"], department=info["department"],
        exam_title=info["exam_title"], semester=info["semester"],
    )

    st.subheader("Downloads")
    d1, d2, d3 = st.columns(3)
    d1.download_button(
        "⬇ Main Seating Plan",
        data=sirt_seating_path.read_bytes(),
        file_name="SIRT_Main_Seating.xlsx", mime=XLSX_MIME,
        use_container_width=True,
    )
    d2.download_button(
        "⬇ Debarred List",
        data=sirt_debarred_path.read_bytes(),
        file_name="SIRT_Debarred_List.xlsx", mime=XLSX_MIME,
        use_container_width=True,
    )
    d3.download_button(
        "⬇ Attendance Sheets",
        data=sirt_att_path.read_bytes(),
        file_name="SIRT_Attendance_Sheets.xlsx", mime=XLSX_MIME,
        use_container_width=True,
    )

    # Combined all-in-one workbook
    combined_bytes = _merge_workbooks(
        [sirt_seating_path, sirt_debarred_path, sirt_att_path],
        ["Seating", "Debarred", "Attendance"],
    )
    st.download_button(
        "⬇ Download All in One File",
        data=combined_bytes,
        file_name="SIRT_All_Reports.xlsx",
        mime=XLSX_MIME,
        use_container_width=True,
        type="primary",
    )


def main() -> None:
    st.set_page_config(
        page_title="SIRT Exam Seating System",
        page_icon="🪑",
        layout="wide",
    )
    st.title("SIRT — Automatic Exam Seating Arrangement")

    info = _sidebar_institute_info()
    tab_generate(info)


if __name__ == "__main__":
    main()
