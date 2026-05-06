"""
Extract students and rooms from '1ST mid sem 6th sem RESULT Jan-jun 2026.xlsx'
and write to output_from_your_file/students.xlsx and rooms.xlsx
"""
from __future__ import annotations
from pathlib import Path
import json
import openpyxl
import pandas as pd

from convert_result_to_students import extract_timetable_from_workbook


INPUT_FILE = "1ST mid sem 6th sem RESULT Jan-jun 2026.xlsx"
OUTPUT_DIR = Path("output_from_your_file")
OUTPUT_DIR.mkdir(exist_ok=True)

SHEET_MAP = {"A": "6th A", "B": "6th B", "C": "6th C", "D": "6th D", "E": "6th E"}
DEFAULT_PREFIX = "0133CI231"
DEFAULT_SUFFIX_LEN = 3
ROOM_CAPACITY = 40  # default capacity per room


def _split_roll(token: str) -> tuple[str, str]:
    """
    Split a full SIRT roll number into (prefix, serial).

    SIRT format examples:
      0133CI231002  →  prefix='0133CI231', serial='002'  (regular)
      0133CI243D15  →  prefix='0133CI243D', serial='15'  (detained)

    Rule: find the last alpha character; everything after it is
    '<year_digits><serial_digits>'. If there are more than 3 trailing
    digits (i.e., year+serial combined), the first 3 form the year
    code and the rest is the serial.
    """
    last_alpha = -1
    for i, c in enumerate(token):
        if c.isalpha():
            last_alpha = i

    after = token[last_alpha + 1:]  # digits after last alpha char
    if len(after) > 3:
        # e.g. "231002" → year="231", serial="002"
        year = after[:3]
        serial = after[3:]
        return token[: last_alpha + 1] + year, serial
    else:
        return token[: last_alpha + 1], after


def expand_rolls(compressed: str, default_prefix: str = DEFAULT_PREFIX, default_suffix_len: int = DEFAULT_SUFFIX_LEN) -> list[str]:
    """Expand compressed roll number string like '0133CI231002,08,14' into full roll numbers."""
    text = compressed.replace("\n", ",").replace(" ", "")
    tokens = [t.strip() for t in text.split(",") if t.strip()]
    results: list[str] = []
    prefix = default_prefix
    suf_len = default_suffix_len
    for token in tokens:
        if any(c.isalpha() for c in token):
            # Full roll number — split correctly using SIRT format rules
            prefix, serial = _split_roll(token)
            suf_len = len(serial)
            results.append(token)
        else:
            padded = token.zfill(suf_len)
            results.append(prefix + padded)
    return results


def read_section_attendance(wb: openpyxl.Workbook, sheet_name: str) -> list[float]:
    """Read attendance percentages in serial-number order from a section sheet."""
    ws = wb[sheet_name]
    att: list[float] = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        sno, _, _, per = row[0], row[1], row[2], row[3]
        if sno is not None and isinstance(sno, (int, float)) and per is not None:
            att.append(float(per))
    return att


def main() -> None:
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    # ── Parse Main Seating sheet ─────────────────────────────────────────────
    # section -> [(roll_list, room)]
    section_data: dict[str, list[tuple[list[str], str]]] = {}
    for row in wb["Main Seating"].iter_rows(min_row=4, values_only=True):
        if not (row[0] and isinstance(row[0], (int, float))):
            continue
        section = str(row[3]).strip() if row[3] else ""
        rolls_raw = str(row[4]).strip() if row[4] else ""
        room = str(row[6]).strip() if row[6] and str(row[6]).strip() != "None" else ""
        if not rolls_raw or not section:
            continue
        rolls = expand_rolls(rolls_raw)
        if section not in section_data:
            section_data[section] = []
        section_data[section].append((rolls, room))

    # ── Build students dataframe ─────────────────────────────────────────────
    all_students: list[dict] = []
    for section, groups in sorted(section_data.items()):
        sheet_name = SHEET_MAP.get(section)
        if not sheet_name or sheet_name not in wb.sheetnames:
            print(f"  [WARN] No attendance sheet for section {section}, skipping.")
            continue

        # Combine all roll-number lists for this section and sort numerically
        # Regular rolls (0133CI231xxx) sort first, detained (0133CI243Dxx) last
        all_rolls: list[str] = []
        for rolls, _ in groups:
            all_rolls.extend(rolls)

        def roll_sort_key(r: str) -> tuple:
            # Extract trailing number for sorting
            i = len(r)
            while i > 0 and r[i - 1].isdigit():
                i -= 1
            prefix = r[:i]
            num = int(r[i:]) if r[i:].isdigit() else 0
            return (prefix, num)

        regular = sorted([r for r in all_rolls if "243D" not in r and "243d" not in r], key=roll_sort_key)
        detained = sorted([r for r in all_rolls if "243D" in r or "243d" in r], key=roll_sort_key)
        ordered_rolls = regular + detained

        att_list = read_section_attendance(wb, sheet_name)

        if len(ordered_rolls) != len(att_list):
            print(
                f"  [WARN] Section {section}: {len(ordered_rolls)} roll numbers "
                f"vs {len(att_list)} attendance rows — will zip to minimum."
            )

        for roll, att in zip(ordered_rolls, att_list):
            all_students.append(
                {
                    "Roll Number": roll,
                    "Name": roll,  # names not in file; use roll as placeholder
                    "Attendance Percentage": round(att, 2),
                    "Branch": "CSIT",
                    "Section": section,
                }
            )

    students_df = pd.DataFrame(all_students)
    students_path = OUTPUT_DIR / "students.xlsx"
    students_df.to_excel(students_path, index=False)
    print(f"Saved {len(students_df)} students → {students_path}")

    # ── Build rooms dataframe ────────────────────────────────────────────────
    rooms: dict[str, int] = {}
    for groups in section_data.values():
        for rolls, room in groups:
            if room:
                rooms[room] = ROOM_CAPACITY

    rooms_df = pd.DataFrame(
        [{"Room Number": r, "Capacity": c} for r, c in sorted(rooms.items())]
    )
    rooms_path = OUTPUT_DIR / "rooms.xlsx"
    rooms_df.to_excel(rooms_path, index=False)
    print(f"Saved {len(rooms_df)} rooms → {rooms_path}")

    # ── Extract timetable ────────────────────────────────────────────────────
    timetable = extract_timetable_from_workbook(Path(INPUT_FILE))
    timetable_path = OUTPUT_DIR / "timetable.json"
    timetable_path.write_text(json.dumps(timetable, indent=2), encoding="utf-8")
    print(f"Saved {len(timetable)} timetable entries → {timetable_path}")
    for entry in timetable:
        print(f"  {entry['date']}  {entry['subject']}")

    print("\nSection summary:")
    for sec in sorted(section_data.keys()):
        sheet_name = SHEET_MAP.get(sec, "?")
        rolls_count = sum(len(r) for r, _ in section_data[sec])
        print(f"  Section {sec} ({sheet_name}): {rolls_count} rolls")


if __name__ == "__main__":
    main()
