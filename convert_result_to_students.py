from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path

import openpyxl
import pandas as pd


ROLL_ALIASES = {
    "rollnumber",
    "rollno",
    "rollno.",
    "enrollno",
    "boardenrollno",
    "enrollmentno",
    "enrollmentnumber",
}
NAME_ALIASES = {"name", "studentname", "studentname."}
ATTENDANCE_ALIASES = {
    "attendancepercentage",
    "attendance",
    "percentage",
    "%percentage",
    "percent",
    "totalper",
    "per",
}

ROOM_ALIASES = {"classroom", "room", "roomno", "roomnumber", "classroomno"}
CAPACITY_ALIASES = {"noofstudentsappeared", "studentsappeared", "capacity", "appeared", "noofstud"}


def _norm(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum() or ch == "%")


def _find_col(columns: list[object], aliases: set[str]) -> str | None:
    for col in columns:
        if _norm(col) in aliases:
            return str(col)
    return None


def _parse_section_branch(sheet_name: str) -> tuple[str, str]:
    """Parse Section and Branch from sheet names like '6th A', '6th B', 'CSIT-A', etc."""
    name = sheet_name.strip()
    # Match trailing single uppercase letter as section, e.g. "6th A" -> A
    m = re.search(r"\b([A-Z])\s*$", name)
    section = m.group(1) if m else ""
    branch_part = re.sub(r"\s*[-_]?\s*[A-Z]\s*$", "", name).strip(" -_")
    branch = branch_part if branch_part else "CSIT"
    return branch, section


def _extract_from_sheet(path: Path, sheet: str) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    branch, section = _parse_section_branch(sheet)

    for header_idx in range(min(30, len(raw))):
        header = [str(x).strip() for x in raw.iloc[header_idx].tolist()]
        data = raw.iloc[header_idx + 1 :].copy()
        data.columns = header
        data = data.dropna(how="all")

        roll_col = _find_col(list(data.columns), ROLL_ALIASES)
        name_col = _find_col(list(data.columns), NAME_ALIASES)
        att_col = _find_col(list(data.columns), ATTENDANCE_ALIASES)

        if roll_col and name_col and att_col:
            out = data[[roll_col, name_col, att_col]].copy()
            out.columns = ["Roll Number", "Name", "Attendance Percentage"]
            out["Roll Number"] = out["Roll Number"].astype(str).str.strip()
            out["Name"] = out["Name"].astype(str).str.strip()
            out["Attendance Percentage"] = pd.to_numeric(
                out["Attendance Percentage"].astype(str).str.replace("%", "", regex=False),
                errors="coerce",
            )
            out = out.dropna(subset=["Attendance Percentage"])
            out = out[(out["Roll Number"] != "") & (out["Name"] != "")]
            if not out.empty:
                out["Branch"] = branch
                out["Section"] = section
                return out.reset_index(drop=True)

    return pd.DataFrame(columns=["Roll Number", "Name", "Attendance Percentage", "Branch", "Section"])


def extract_rooms_from_workbook(path: Path) -> pd.DataFrame:
    """Try to extract Room Number and Capacity from any sheet in the workbook."""
    xl = pd.ExcelFile(path)
    for sheet in xl.sheet_names:
        raw = pd.read_excel(path, sheet_name=sheet, header=None)
        for header_idx in range(min(30, len(raw))):
            header = [str(x).strip() for x in raw.iloc[header_idx].tolist()]
            data = raw.iloc[header_idx + 1 :].copy()
            data.columns = header
            data = data.dropna(how="all")
            room_col = _find_col(list(data.columns), ROOM_ALIASES)
            cap_col = _find_col(list(data.columns), CAPACITY_ALIASES)
            if room_col and cap_col:
                rooms = data[[room_col, cap_col]].copy()
                rooms.columns = ["Room Number", "Capacity"]
                rooms["Room Number"] = rooms["Room Number"].astype(str).str.strip()
                rooms["Capacity"] = pd.to_numeric(rooms["Capacity"], errors="coerce")
                rooms = rooms.dropna(subset=["Capacity"])
                rooms = rooms[rooms["Room Number"].str.len() > 0]
                rooms["Capacity"] = rooms["Capacity"].astype(int)
                rooms = rooms[rooms["Capacity"] > 0]
                rooms = rooms.groupby("Room Number", as_index=False)["Capacity"].max()
                if not rooms.empty:
                    return rooms.reset_index(drop=True)
    return pd.DataFrame(columns=["Room Number", "Capacity"])


def extract_timetable_from_workbook(path: Path) -> list[dict]:
    """
    Extract the exam timetable (subject + date pairs) from any section sheet.

    Looks for the pattern where:
      - A header row contains datetime objects (exam dates) in columns after the
        fixed student columns (S.NO, Board EnrollNo, Student Name, Total Per).
      - The very next row contains subject codes / names in those same columns.

    Returns a list of dicts: [{"subject": "CSIT-601 SE", "date": "06-Apr-26"}, ...]
    """
    xl = pd.ExcelFile(path)
    # Try each sheet; use the first one that yields timetable data
    for sheet in xl.sheet_names:
        if "debarred" in sheet.lower():
            continue
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        except Exception:
            continue

        for i, row in enumerate(rows):
            # Find a row that has datetime values starting from column index 4
            date_cols = [
                (col_idx, val)
                for col_idx, val in enumerate(row)
                if isinstance(val, dt.datetime) and col_idx >= 4
            ]
            if not date_cols:
                continue
            # The next row should have subject names
            if i + 1 >= len(rows):
                continue
            subj_row = rows[i + 1]
            timetable: list[dict] = []
            for col_idx, date_val in date_cols:
                subj = subj_row[col_idx] if col_idx < len(subj_row) else None
                if not subj:
                    continue
                subj_clean = str(subj).replace("\n", " ").strip()
                date_str = date_val.strftime("%d-%b-%y")
                timetable.append({"subject": subj_clean, "date": date_str})
            if timetable:
                return timetable

    return []


def convert_result_file(
    path: Path,
    sheets: list[str] | None = None,
    skip_debarred: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Convert a result-style workbook into (students_df, rooms_df).
    Returns students with Branch/Section columns preserved.
    """
    xl = pd.ExcelFile(path)
    target_sheets = sheets if sheets else xl.sheet_names

    if skip_debarred:
        target_sheets = [s for s in target_sheets if "debarred" not in s.lower()]

    frames: list[pd.DataFrame] = []
    for sheet in target_sheets:
        if sheet not in xl.sheet_names:
            continue
        extracted = _extract_from_sheet(path, sheet)
        if not extracted.empty:
            frames.append(extracted)

    students = pd.DataFrame(columns=["Roll Number", "Name", "Attendance Percentage", "Branch", "Section"])
    if frames:
        students = pd.concat(frames, ignore_index=True)
        students = students.drop_duplicates(subset=["Roll Number"], keep="first")
        students = students.sort_values(by=["Roll Number"], kind="mergesort").reset_index(drop=True)

    rooms = extract_rooms_from_workbook(path)
    return students, rooms


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert result-style workbook into students input format"
    )
    parser.add_argument("--input", required=True, help="Path to source Excel file")
    parser.add_argument(
        "--output",
        default="students_from_result.xlsx",
        help="Output students Excel path",
    )
    parser.add_argument(
        "--rooms-output",
        default=None,
        help="Output rooms Excel path (optional)",
    )
    parser.add_argument(
        "--sheets",
        nargs="*",
        default=None,
        help="Optional specific sheet names to parse (default: all non-debarred)",
    )
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"Input file not found: {src}")
        return 1

    students, rooms = convert_result_file(src, sheets=args.sheets)

    if students.empty:
        print("No valid student table found. Ensure sheets contain roll/name/percentage-like columns.")
        return 1

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    students.to_excel(out_path, index=False)
    print(f"Created {out_path} with {len(students)} students")

    if args.rooms_output and not rooms.empty:
        rooms_path = Path(args.rooms_output)
        rooms_path.parent.mkdir(parents=True, exist_ok=True)
        rooms.to_excel(rooms_path, index=False)
        print(f"Created {rooms_path} with {len(rooms)} rooms")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
